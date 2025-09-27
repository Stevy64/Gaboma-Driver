from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import logout
from django.contrib.auth.models import User, Group
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Avg, Q
from django.db import models
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime, date, timedelta
from drivers.models import Chauffeur, AssignationSuperviseur
from activities.models import Activite, Recette, Panne, PriseCles, RemiseCles, DemandeModification
from functools import wraps


def supervisor_required(view_func):
    """
    Décorateur personnalisé pour vérifier les privilèges de superviseur
    Accepte :
    - Superusers
    - Utilisateurs avec is_staff = True
    - Utilisateurs du groupe 'Superviseurs'
    """
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.error(request, 'Vous devez être connecté pour accéder à cette page.')
            return redirect('drivers:login_superviseur')
        
        # Vérifier les privilèges superviseur
        is_supervisor_group = request.user.groups.filter(name='Superviseurs').exists()
        is_staff_user = request.user.is_staff
        is_superuser = request.user.is_superuser
        
        # Vérifier aussi si c'est un chauffeur avec is_staff
        is_chauffeur_with_staff = False
        try:
            from drivers.models import Chauffeur
            chauffeur = Chauffeur.objects.get(user=request.user)
            is_chauffeur_with_staff = chauffeur.actif and request.user.is_staff
        except Chauffeur.DoesNotExist:
            pass
        
        if not (is_superuser or is_staff_user or is_supervisor_group or is_chauffeur_with_staff):
            messages.error(request, 'Vous n\'avez pas les privilèges nécessaires pour accéder à cette page.')
            return redirect('drivers:index')
        
        return view_func(request, *args, **kwargs)
    return wrapper


def get_chauffeurs_for_user(user):
    """
    Récupère les chauffeurs accessibles selon le type d'utilisateur
    
    Args:
        user: Utilisateur connecté
        
    Returns:
        QuerySet: Chauffeurs accessibles
    """
    if user.is_superuser or user.is_staff:
        # Super admin ou superviseur avec is_staff : accès à tous les chauffeurs
        return Chauffeur.objects.all()
    elif user.groups.filter(name='Superviseurs').exists():
        # Superviseur du groupe (sans is_staff) : seulement ses chauffeurs assignés
        return AssignationSuperviseur.get_chauffeurs_assignes(user)
    else:
        # Autres utilisateurs : aucun accès
        return Chauffeur.objects.none()


def get_activites_for_user(user, model_class, **filters):
    """
    Récupère les activités accessibles selon le type d'utilisateur
    
    Args:
        user: Utilisateur connecté
        model_class: Classe du modèle d'activité
        **filters: Filtres additionnels
        
    Returns:
        QuerySet: Activités accessibles
    """
    chauffeurs = get_chauffeurs_for_user(user)
    if chauffeurs.exists():
        return model_class.objects.filter(chauffeur__in=chauffeurs, **filters)
    else:
        return model_class.objects.none()

# Import conditionnel d'openpyxl pour éviter les erreurs si le module n'est pas installé
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def logout_admin(request):
    """
    Vue de déconnexion pour les administrateurs
    
    Cette vue déconnecte l'utilisateur admin et le redirige vers la page d'accueil
    avec un message de confirmation.
    
    Args:
        request: Objet HttpRequest de l'utilisateur
        
    Returns:
        HttpResponse: Redirection vers la page d'accueil
    """
    logout(request)
    messages.success(request, 'Vous avez été déconnecté avec succès.')
    return redirect('drivers:index')

def dashboard_superviseur(request):
    """
    Dashboard spécifique pour les superviseurs (non super admin)
    
    Affiche les mêmes statistiques que le dashboard admin mais avec:
    - Navigation adaptée aux superviseurs
    - Bouton Admin conditionnel selon les privilèges staff
    - Accès limité aux chauffeurs assignés
    """
    # Vérifier que l'utilisateur a les privilèges de superviseur
    # Soit membre du groupe 'Superviseurs', soit chauffeur avec is_staff = True
    is_supervisor_group = request.user.groups.filter(name='Superviseurs').exists()
    is_chauffeur_with_staff = False
    
    try:
        from drivers.models import Chauffeur
        chauffeur = Chauffeur.objects.get(user=request.user)
        is_chauffeur_with_staff = chauffeur.actif and request.user.is_staff
    except Chauffeur.DoesNotExist:
        pass
    
    if not (is_supervisor_group or is_chauffeur_with_staff):
        messages.error(request, 'Accès refusé. Vous devez être superviseur ou chauffeur avec statut équipe pour accéder à cet espace.')
        return redirect('drivers:index')
    
    # Si c'est un superuser ou utilisateur avec is_staff, rediriger vers le dashboard admin complet
    if request.user.is_superuser or request.user.is_staff:
        return redirect('admin_dashboard:dashboard_admin')
    
    # Récupérer les chauffeurs accessibles selon le type d'utilisateur
    # La fonction get_chauffeurs_for_user gère déjà toute la logique de filtrage
    chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
    
    # Statistiques générales (identiques au dashboard admin mais filtrées)
    total_chauffeurs = chauffeurs_accessibles.filter(actif=True).count()
    
    # Activités du jour
    prises_aujourdhui = get_activites_for_user(request.user, PriseCles, date=date.today()).count()
    remises_aujourdhui = get_activites_for_user(request.user, RemiseCles, date=date.today()).count()
    total_activites_aujourdhui = prises_aujourdhui + remises_aujourdhui
    
    # Recettes
    recettes_aujourdhui = get_activites_for_user(request.user, RemiseCles, date=date.today()).aggregate(total=Sum('recette_realisee'))['total'] or 0
    recettes_semaine = get_activites_for_user(request.user, RemiseCles, date__gte=date.today() - timedelta(days=7)).aggregate(total=Sum('recette_realisee'))['total'] or 0
    recettes_mois = get_activites_for_user(request.user, RemiseCles, date__gte=date.today().replace(day=1)).aggregate(total=Sum('recette_realisee'))['total'] or 0
    
    # Pannes
    pannes_en_cours = get_activites_for_user(request.user, Panne, statut__in=['signalee', 'en_cours']).count()
    pannes_critiques = get_activites_for_user(request.user, Panne, severite='critique', statut__in=['signalee', 'en_cours']).count()
    
    # Activités récentes (limitées)
    activites_recentes = []
    prises_recentes = get_activites_for_user(request.user, PriseCles).order_by('-date', '-heure_prise')[:5]
    remises_recentes = get_activites_for_user(request.user, RemiseCles).order_by('-date', '-heure_remise')[:5]
    
    for prise in prises_recentes:
        activites_recentes.append({
            'type': 'prise',
            'chauffeur': prise.chauffeur,
            'date': prise.date,
            'heure': prise.heure_prise,
            'details': f"Prise de clés - Objectif: {prise.objectif_recette} FCFA"
        })
    
    for remise in remises_recentes:
        activites_recentes.append({
            'type': 'remise',
            'chauffeur': remise.chauffeur,
            'date': remise.date,
            'heure': remise.heure_remise,
            'details': f"Remise de clés - {remise.recette_realisee} FCFA"
        })
    
    # Trier par date et heure décroissante
    activites_recentes = sorted(activites_recentes, key=lambda x: (x['date'], x['heure']), reverse=True)[:10]
    
    # Pannes récentes
    pannes_recentes = get_activites_for_user(request.user, Panne).order_by('-date_creation')[:5]
    
    # Demandes récentes (dernières 24h)
    hier = timezone.now() - timedelta(hours=24)
    demandes_recentes = DemandeModification.objects.select_related('chauffeur').filter(
        chauffeur__in=chauffeurs_accessibles,
        date_creation__gte=hier
    ).order_by('-date_creation')[:10]
    
    # Alertes spéciales pour superviseur
    demandes_en_attente = DemandeModification.objects.filter(
        chauffeur__in=chauffeurs_accessibles,
        statut='en_attente'
    ).count()
    
    # Activités très récentes (aujourd'hui)
    prises_recentes_aujourdhui = get_activites_for_user(request.user, PriseCles, date=date.today()).count()
    remises_recentes_aujourdhui = get_activites_for_user(request.user, RemiseCles, date=date.today()).count()
    
    # Vérifier si l'utilisateur a le privilège "Statut équipe" (is_staff)
    has_staff_privilege = request.user.is_staff
    
    context = {
        'total_chauffeurs': total_chauffeurs,
        'total_activites_aujourdhui': total_activites_aujourdhui,
        'prises_aujourdhui': prises_aujourdhui,
        'remises_aujourdhui': remises_aujourdhui,
        'recettes_aujourdhui': recettes_aujourdhui,
        'recettes_semaine': recettes_semaine,
        'recettes_mois': recettes_mois,
        'pannes_en_cours': pannes_en_cours,
        'pannes_critiques': pannes_critiques,
        'activites_recentes': activites_recentes,
        'pannes_recentes': pannes_recentes,
        'is_supervisor': True,
        'has_staff_privilege': has_staff_privilege,
        'is_supervisor_group': is_supervisor_group,
        'is_chauffeur_with_staff': is_chauffeur_with_staff,
        'demandes_recentes': demandes_recentes,
        'demandes_en_attente': demandes_en_attente,
        'prises_recentes_aujourdhui': prises_recentes_aujourdhui,
        'remises_recentes_aujourdhui': remises_recentes_aujourdhui,
    }
    
    return render(request, 'admin_dashboard/dashboard_superviseur.html', context)


@supervisor_required
def dashboard_admin(request):
    # Vérifier si l'utilisateur est un superviseur simple (pas super admin ni is_staff)
    is_supervisor = request.user.groups.filter(name='Superviseurs').exists() and not request.user.is_superuser and not request.user.is_staff
    """
    Tableau de bord administrateur avec statistiques en temps réel
    
    Affiche les statistiques générales de l'application :
    - Nombre de chauffeurs actifs
    - Activités du jour (prises et remises)
    - Recettes en FCFA (jour, semaine, mois)
    - Pannes en cours et critiques
    - Top chauffeurs du mois
    - Activités et pannes récentes
    """
    # Récupérer les chauffeurs accessibles selon le type d'utilisateur
    chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
    
    # Statistiques générales
    total_chauffeurs = chauffeurs_accessibles.filter(actif=True).count()
    
    # Activités du jour (filtrées par chauffeurs accessibles)
    prises_aujourdhui = get_activites_for_user(request.user, PriseCles, date=date.today()).count()
    remises_aujourdhui = get_activites_for_user(request.user, RemiseCles, date=date.today()).count()
    total_activites_aujourdhui = prises_aujourdhui + remises_aujourdhui
    
    # Recettes en FCFA (filtrées par chauffeurs accessibles)
    recettes_aujourdhui = get_activites_for_user(request.user, RemiseCles, date=date.today()).aggregate(total=Sum('recette_realisee'))['total'] or 0
    
    recettes_semaine = get_activites_for_user(request.user, RemiseCles, date__gte=date.today() - timedelta(days=7)).aggregate(total=Sum('recette_realisee'))['total'] or 0
    
    recettes_mois = get_activites_for_user(request.user, RemiseCles, date__gte=date.today().replace(day=1)).aggregate(total=Sum('recette_realisee'))['total'] or 0
    
    # Pannes (filtrées par chauffeurs accessibles)
    pannes_en_cours = get_activites_for_user(request.user, Panne, statut__in=['signalee', 'en_cours']).count()
    pannes_critiques = get_activites_for_user(request.user, Panne, severite='critique').count()
    
    
    # Activités récentes (prises et remises) - filtrées par chauffeurs accessibles
    prises_recentes = get_activites_for_user(request.user, PriseCles).select_related('chauffeur').order_by('-date', '-heure_prise')
    remises_recentes = get_activites_for_user(request.user, RemiseCles).select_related('chauffeur').order_by('-date', '-heure_remise')
    
    # Combiner les activités récentes pour l'affichage
    activites_recentes = []
    for prise in prises_recentes:
        activites_recentes.append({
            'chauffeur': prise.chauffeur,
            'type_activite': 'prise',
            'date_heure': f"{prise.date} {prise.heure_prise}",
            'date': prise.date,
            'heure': prise.heure_prise
        })
    for remise in remises_recentes:
        activites_recentes.append({
            'chauffeur': remise.chauffeur,
            'type_activite': 'remise',
            'date_heure': f"{remise.date} {remise.heure_remise}",
            'date': remise.date,
            'heure': remise.heure_remise
        })
    
    # Trier par date décroissante
    activites_recentes.sort(key=lambda x: (x['date'], x['heure']), reverse=True)
    
    # Pagination des activités récentes
    from django.core.paginator import Paginator
    activites_paginator = Paginator(activites_recentes, 10)  # 10 activités par page
    activites_page = request.GET.get('activites_page')
    activites_obj = activites_paginator.get_page(activites_page)
    
    # Demandes de modification en attente
    demandes_en_attente = DemandeModification.objects.filter(statut='en_attente').count()
    
    # Pannes récentes pour affichage avec pagination
    pannes_recentes = Panne.objects.select_related('chauffeur').order_by('-date_creation')
    pannes_paginator = Paginator(pannes_recentes, 10)  # 10 pannes par page
    pannes_page = request.GET.get('pannes_page')
    pannes_obj = pannes_paginator.get_page(pannes_page)
    
    context = {
        'total_chauffeurs': total_chauffeurs,
        'total_activites_aujourdhui': total_activites_aujourdhui,
        'prises_aujourdhui': prises_aujourdhui,
        'remises_aujourdhui': remises_aujourdhui,
        'recettes_aujourdhui': recettes_aujourdhui,
        'recettes_semaine': recettes_semaine,
        'recettes_mois': recettes_mois,
        'pannes_en_cours': pannes_en_cours,
        'pannes_critiques': pannes_critiques,
        'demandes_en_attente': demandes_en_attente,
        'activites_recentes': activites_obj,
        'activites_page_obj': activites_obj,
        'pannes_recentes': pannes_obj,
        'pannes_page_obj': pannes_obj,
    }
    
    return render(request, 'admin_dashboard/dashboard.html', context)


@supervisor_required
def liste_chauffeurs(request):
    """Liste des chauffeurs avec pagination"""
    from django.core.paginator import Paginator
    
    # Filtrer les chauffeurs selon le type d'utilisateur
    chauffeurs = get_chauffeurs_for_user(request.user).order_by('nom', 'prenom')
    
    # Pagination
    paginator = Paginator(chauffeurs, 10)  # 10 chauffeurs par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Déterminer les permissions de l'utilisateur
    can_modify_chauffeurs = request.user.is_superuser or request.user.is_staff
    is_simple_supervisor = (request.user.groups.filter(name='Superviseurs').exists() and 
                           not request.user.is_superuser and 
                           not request.user.is_staff)
    
    context = {
        'chauffeurs': page_obj,
        'page_obj': page_obj,
        'can_modify_chauffeurs': can_modify_chauffeurs,
        'is_simple_supervisor': is_simple_supervisor,
    }
    
    return render(request, 'admin_dashboard/liste_chauffeurs.html', context)


@supervisor_required
def statistiques_recettes(request):
    """
    Statistiques des recettes avec filtres par période et chauffeur
    
    Affiche les recettes en FCFA pour différentes périodes :
    - Par chauffeur avec nombre de jours travaillés
    - Par jour avec évolution temporelle
    - Totaux par période (jour, semaine, mois, année)
    - Graphiques et disques statistiques enrichis
    """
    # Filtres de période et chauffeur
    periode = request.GET.get('periode', 'mois')
    chauffeur_id = request.GET.get('chauffeur', '')
    
    # Définition des périodes
    if periode == 'jour':
        date_debut = date.today()
        date_fin = date.today()
    elif periode == 'semaine':
        # Semaine courante (lundi à dimanche)
        today = date.today()
        date_debut = today - timedelta(days=today.weekday())
        date_fin = date_debut + timedelta(days=6)
    elif periode == 'mois':
        date_debut = date.today().replace(day=1)
        date_fin = date.today()
    elif periode == 'annee':
        date_debut = date.today().replace(month=1, day=1)
        date_fin = date.today()
    else:
        date_debut = date.today().replace(day=1)
        date_fin = date.today()
    
    # Filtre par chauffeur si spécifié
    chauffeur_filter = {}
    if chauffeur_id:
        try:
            chauffeur = Chauffeur.objects.get(id=chauffeur_id)
            chauffeur_filter = {'chauffeur': chauffeur}
        except Chauffeur.DoesNotExist:
            chauffeur = None
    else:
        chauffeur = None
    
    # Recettes par chauffeur (filtrées selon les permissions utilisateur)
    chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
    
    recettes_chauffeurs = chauffeurs_accessibles.filter(
        remisecles__date__gte=date_debut,
        remisecles__date__lte=date_fin
    ).annotate(
        total_recettes=Sum('remisecles__recette_realisee'),
        nb_jours=Count('remisecles', distinct=True),
        moyenne_journaliere=Avg('remisecles__recette_realisee')
    ).filter(total_recettes__gt=0).order_by('-total_recettes')
    
    # Recettes par jour (utilisant les nouveaux modèles)
    recettes_par_jour = RemiseCles.objects.filter(
        date__gte=date_debut,
        date__lte=date_fin,
        **chauffeur_filter
    ).values('date').annotate(
        total=Sum('recette_realisee'),
        nb_chauffeurs=Count('chauffeur', distinct=True)
    ).order_by('date')
    
    # Calcul des totaux
    recette_totale = RemiseCles.objects.filter(
        date__gte=date_debut,
        date__lte=date_fin,
        **chauffeur_filter
    ).aggregate(Sum('recette_realisee'))['recette_realisee__sum'] or 0
    
    # Statistiques supplémentaires adaptées aux filtres
    if chauffeur:
        # Si un chauffeur spécifique est sélectionné
        nombre_chauffeurs_actifs = 1
        moyenne_par_chauffeur = recette_totale
        nom_chauffeur = chauffeur.nom_complet
    else:
        # Si tous les chauffeurs sont sélectionnés
        nombre_chauffeurs_actifs = chauffeurs_accessibles.filter(actif=True).count()
        moyenne_par_chauffeur = recette_totale / nombre_chauffeurs_actifs if nombre_chauffeurs_actifs > 0 else 0
        nom_chauffeur = "Chauffeurs Actifs"
    
    # Données pour les graphiques
    # 1. Données pour le graphique en barres (recettes par jour)
    chart_data_daily = []
    for recette_jour in recettes_par_jour:
        chart_data_daily.append({
            'date': recette_jour['date'].strftime('%d/%m'),
            'recette': float(recette_jour['total']),
            'nb_chauffeurs': recette_jour['nb_chauffeurs']
        })
    
    # 2. Données pour le graphique en secteurs (répartition par chauffeur)
    chart_data_chauffeurs = []
    for chauffeur_data in recettes_chauffeurs:
        chart_data_chauffeurs.append({
            'nom': chauffeur_data.nom_complet,
            'recette': float(chauffeur_data.total_recettes),
            'nb_jours': chauffeur_data.nb_jours,
            'moyenne': float(chauffeur_data.moyenne_journaliere or 0)
        })
    
    # 3. Statistiques pour les disques de performance
    # Calcul de la performance moyenne
    performances_chauffeurs = []
    for chauffeur_data in recettes_chauffeurs:
        # Récupérer les objectifs pour ce chauffeur sur la période
        objectifs_chauffeur = PriseCles.objects.filter(
            chauffeur=chauffeur_data,
            date__gte=date_debut,
            date__lte=date_fin
        ).aggregate(Sum('objectif_recette'))['objectif_recette__sum'] or 0
        
        performance = (chauffeur_data.total_recettes / objectifs_chauffeur * 100) if objectifs_chauffeur > 0 else 0
        performances_chauffeurs.append({
            'chauffeur': chauffeur_data.nom_complet,
            'performance': performance,
            'recette': chauffeur_data.total_recettes,
            'objectif': objectifs_chauffeur
        })
    
    # Performance moyenne globale
    performance_moyenne = sum(p['performance'] for p in performances_chauffeurs) / len(performances_chauffeurs) if performances_chauffeurs else 0
    
    # 4. Données pour l'évolution temporelle (si période > jour)
    evolution_data = []
    if periode in ['semaine', 'mois', 'annee']:
        # Calculer les totaux par sous-période
        if periode == 'semaine':
            # Par jour de la semaine
            for i in range(7):
                jour_date = date_debut + timedelta(days=i)
                if jour_date <= date_fin:
                    recette_jour = RemiseCles.objects.filter(
                        date=jour_date,
                        **chauffeur_filter
                    ).aggregate(Sum('recette_realisee'))['recette_realisee__sum'] or 0
                    evolution_data.append({
                        'periode': jour_date.strftime('%A'),
                        'recette': float(recette_jour)
                    })
        elif periode == 'mois':
            # Par semaine du mois
            semaine_actuelle = 1
            recette_semaine = 0
            for recette_jour in recettes_par_jour:
                semaine_jour = recette_jour['date'].isocalendar()[1]
                if semaine_jour != semaine_actuelle:
                    evolution_data.append({
                        'periode': f'Semaine {semaine_actuelle}',
                        'recette': float(recette_semaine)
                    })
                    semaine_actuelle = semaine_jour
                    recette_semaine = recette_jour['total']
                else:
                    recette_semaine += recette_jour['total']
            # Ajouter la dernière semaine
            if recette_semaine > 0:
                evolution_data.append({
                    'periode': f'Semaine {semaine_actuelle}',
                    'recette': float(recette_semaine)
                })
        elif periode == 'annee':
            # Par mois de l'année
            for mois in range(1, 13):
                recette_mois = RemiseCles.objects.filter(
                    date__year=date.today().year,
                    date__month=mois,
                    **chauffeur_filter
                ).aggregate(Sum('recette_realisee'))['recette_realisee__sum'] or 0
                evolution_data.append({
                    'periode': f'{mois:02d}',
                    'recette': float(recette_mois)
                })
    
    # Liste des chauffeurs pour le filtre (filtrée selon les permissions)
    tous_chauffeurs = chauffeurs_accessibles.filter(actif=True).order_by('nom', 'prenom')
    
    context = {
        'periode': periode,
        'chauffeur_selectionne': chauffeur,
        'chauffeur_id': chauffeur_id,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'recettes_chauffeurs': recettes_chauffeurs,
        'recettes_par_jour': recettes_par_jour,
        'recette_totale': recette_totale,
        'nombre_chauffeurs_actifs': nombre_chauffeurs_actifs,
        'moyenne_par_chauffeur': moyenne_par_chauffeur,
        'nom_chauffeur': nom_chauffeur,
        'recette_jour': recette_totale if periode == 'jour' else 0,
        'recette_semaine': recette_totale if periode == 'semaine' else 0,
        'recette_mois': recette_totale if periode == 'mois' else 0,
        'recette_annee': recette_totale if periode == 'annee' else 0,
        # Données pour les graphiques
        'chart_data_daily': chart_data_daily,
        'chart_data_chauffeurs': chart_data_chauffeurs,
        'performances_chauffeurs': performances_chauffeurs,
        'performance_moyenne': performance_moyenne,
        'evolution_data': evolution_data,
        'tous_chauffeurs': tous_chauffeurs,
    }
    
    return render(request, 'admin_dashboard/statistiques_recettes.html', context)


@supervisor_required
def calendrier_activites(request):
    """
    Calendrier des activités pour l'administrateur
    
    Utilise la même logique que le calendrier chauffeur mais pour tous les chauffeurs.
    """
    import calendar
    
    # Paramètres de navigation
    today = date.today()
    annee = int(request.GET.get('annee', today.year))
    mois = int(request.GET.get('mois', today.month))
    chauffeur_id = request.GET.get('chauffeur')
    
    # Calculer les dates de début et fin du mois
    mois_debut = date(annee, mois, 1)
    if mois == 12:
        mois_fin = date(annee + 1, 1, 1) - timedelta(days=1)
    else:
        mois_fin = date(annee, mois + 1, 1) - timedelta(days=1)
    
    # Récupération des chauffeurs pour le filtre (filtrée selon les permissions)
    chauffeurs = get_chauffeurs_for_user(request.user).filter(actif=True).order_by('nom', 'prenom')
    
    # Récupération des données du mois
    prises_query = get_activites_for_user(request.user, PriseCles, 
        date__gte=mois_debut,
        date__lte=mois_fin
    ).select_related('chauffeur').order_by('date')
    
    remises_query = get_activites_for_user(request.user, RemiseCles,
        date__gte=mois_debut,
        date__lte=mois_fin
    ).select_related('chauffeur').order_by('date')
    
    # Filtrage par chauffeur si spécifié
    if chauffeur_id:
        prises_query = prises_query.filter(chauffeur_id=chauffeur_id)
        remises_query = remises_query.filter(chauffeur_id=chauffeur_id)
    
    prises_mois = list(prises_query)
    remises_mois = list(remises_query)
    
    # Créer le calendrier du mois (même logique que chauffeur)
    calendrier = creer_calendrier_admin_mensuel(annee, mois, prises_mois, remises_mois)
    
    # Calculer les statistiques du mois
    total_mois = sum(remise.recette_realisee for remise in remises_mois)
    jours_travailles = len(remises_mois)
    moyenne_journaliere = total_mois / jours_travailles if jours_travailles > 0 else 0
    
    # Statistiques par mois de l'année
    stats_par_mois = []
    for m in range(1, 13):
        mois_debut_m = date(annee, m, 1)
        if m == 12:
            mois_fin_m = date(annee + 1, 1, 1) - timedelta(days=1)
        else:
            mois_fin_m = date(annee, m + 1, 1) - timedelta(days=1)
        
        remises_m = get_activites_for_user(request.user, RemiseCles,
            date__gte=mois_debut_m, 
            date__lte=mois_fin_m
        )
        if chauffeur_id:
            remises_m = remises_m.filter(chauffeur_id=chauffeur_id)
            
        total_m = sum(remise.recette_realisee for remise in remises_m)
        jours_m = remises_m.count()
        
        stats_par_mois.append({
            'mois': m,
            'nom_mois': ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun', 
                         'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc'][m-1],
            'total': total_m,
            'jours': jours_m,
            'moyenne': total_m / jours_m if jours_m > 0 else 0,
            'actif': jours_m > 0
        })
    
    context = {
        'annee': annee,
        'mois': mois,
        'mois_debut': mois_debut,
        'mois_fin': mois_fin,
        'calendrier': calendrier,
        'prises_mois': prises_mois,
        'remises_mois': remises_mois,
        'total_mois': total_mois,
        'jours_travailles': jours_travailles,
        'moyenne_journaliere': moyenne_journaliere,
        'stats_par_mois': stats_par_mois,
        'chauffeurs': chauffeurs,
        'chauffeur_selectionne': chauffeur_id,
        'today': today,
    }
    
    return render(request, 'admin_dashboard/calendrier_activites.html', context)


def creer_calendrier_admin_mensuel(annee, mois, prises, remises):
    """
    Créer un calendrier mensuel avec les données d'activité pour l'admin
    
    Version adaptée de la fonction chauffeur pour gérer plusieurs chauffeurs.
    """
    import calendar
    
    # Création du calendrier du mois avec le module calendar
    cal = calendar.monthcalendar(annee, mois)
    
    # Création de dictionnaires pour un accès rapide aux données
    prises_dict = {}
    remises_dict = {}
    
    # Grouper par date pour gérer plusieurs chauffeurs
    for prise in prises:
        if prise.date not in prises_dict:
            prises_dict[prise.date] = []
        prises_dict[prise.date].append(prise)
    
    for remise in remises:
        if remise.date not in remises_dict:
            remises_dict[remise.date] = []
        remises_dict[remise.date].append(remise)
    
    # Construction de la structure de données pour le template
    calendrier_semaines = []
    
    for semaine in cal:
        semaine_data = []
        for jour in semaine:
            if jour == 0:  # Jour vide (hors du mois)
                semaine_data.append(None)
            else:
                # Création de la date du jour
                jour_date = date(annee, mois, jour)
                
                # Récupération des activités pour ce jour
                prises_jour = prises_dict.get(jour_date, [])
                remises_jour = remises_dict.get(jour_date, [])
                
                # Calcul des totaux pour ce jour
                total_recette = sum(remise.recette_realisee for remise in remises_jour)
                total_objectif = sum(prise.objectif_recette for prise in prises_jour)
                
                # Construction des données du jour
                jour_data = {
                    'jour': jour,
                    'date': jour_date,
                    'prises': prises_jour,
                    'remises': remises_jour,
                    'recette': total_recette,
                    'objectif': total_objectif,
                    'actif': len(prises_jour) > 0 or len(remises_jour) > 0,
                    'complet': len(prises_jour) > 0 and len(remises_jour) > 0,
                }
                
                # Calcul du pourcentage de performance
                if jour_data['objectif'] > 0:
                    jour_data['performance'] = int((jour_data['recette'] / jour_data['objectif']) * 100)
                else:
                    jour_data['performance'] = 0
                
                semaine_data.append(jour_data)
        
        calendrier_semaines.append(semaine_data)
    
    return calendrier_semaines


@supervisor_required
def gestion_pannes(request):
    """Gestion des pannes"""
    # Filtrer les pannes selon les chauffeurs accessibles
    chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
    pannes = Panne.objects.select_related('chauffeur').filter(
        chauffeur__in=chauffeurs_accessibles
    ).order_by('-date_creation')
    
    # Filtres
    statut = request.GET.get('statut')
    severite = request.GET.get('severite')
    
    if statut:
        pannes = pannes.filter(resolue=(statut == 'resolue'))
    if severite:
        pannes = pannes.filter(severite=severite)
    
    context = {
        'pannes': pannes,
        'statut_actuel': statut,
        'severite_actuelle': severite,
    }
    
    return render(request, 'admin_dashboard/gestion_pannes.html', context)


# =============================================================================
# GESTION DES ACTIVITÉS - Nouvelles fonctionnalités administrateur
# =============================================================================

@supervisor_required
def gestion_activites(request):
    """
    Vue pour la gestion des activités par chauffeur
    
    Permet aux administrateurs de voir toutes les activités (prises et remises de clés)
    par chauffeur avec possibilité de filtrer et de consulter les détails.
    """
    # Récupération des paramètres de filtrage
    chauffeur_id = request.GET.get('chauffeur')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    type_activite = request.GET.get('type_activite')
    
    # Filtrage des chauffeurs selon les permissions utilisateur
    chauffeurs = get_chauffeurs_for_user(request.user).filter(actif=True).order_by('nom', 'prenom')
    
    # Récupération des activités avec filtres (filtrées selon les permissions)
    prises = get_activites_for_user(request.user, PriseCles)
    remises = get_activites_for_user(request.user, RemiseCles)
    
    # Application des filtres
    if chauffeur_id:
        prises = prises.filter(chauffeur_id=chauffeur_id)
        remises = remises.filter(chauffeur_id=chauffeur_id)
    
    if date_debut:
        try:
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            prises = prises.filter(date__gte=date_debut_obj)
            remises = remises.filter(date__gte=date_debut_obj)
        except ValueError:
            pass
    
    if date_fin:
        try:
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
            prises = prises.filter(date__lte=date_fin_obj)
            remises = remises.filter(date__lte=date_fin_obj)
        except ValueError:
            pass
    
    # Tri par date décroissante
    prises = prises.order_by('-date', '-heure_prise')
    remises = remises.order_by('-date', '-heure_remise')
    
    # Filtrage par type d'activité (pour l'affichage)
    if type_activite == 'prise':
        remises = remises.none()  # Afficher seulement les prises
    elif type_activite == 'remise':
        prises = prises.none()  # Afficher seulement les remises
    
    # Pagination pour les prises
    prises_paginator = Paginator(prises, 10)  # 10 éléments par page
    prises_page = request.GET.get('prises_page')
    prises_obj = prises_paginator.get_page(prises_page)
    
    # Pagination pour les remises
    remises_paginator = Paginator(remises, 10)  # 10 éléments par page
    remises_page = request.GET.get('remises_page')
    remises_obj = remises_paginator.get_page(remises_page)
    
    # Statistiques
    total_prises = prises.count()
    total_remises = remises.count()
    
    # Calcul des recettes totales pour la période
    recettes_totales = remises.aggregate(total=Sum('recette_realisee'))['total'] or 0
    
    context = {
        'chauffeurs': chauffeurs,
        'prises': prises_obj,
        'remises': remises_obj,
        'prises_paginator': prises_paginator,
        'remises_paginator': remises_paginator,
        'total_prises': total_prises,
        'total_remises': total_remises,
        'recettes_totales': recettes_totales,
        'filtres': {
            'chauffeur_id': chauffeur_id,
            'date_debut': date_debut,
            'date_fin': date_fin,
            'type_activite': type_activite,
        },
        'filtres_applied': {
            'chauffeur': chauffeur_id,
            'date_debut': date_debut,
            'date_fin': date_fin,
            'type_activite': type_activite,
        }
    }
    
    return render(request, 'admin_dashboard/gestion_activites.html', context)


@supervisor_required
def activites_chauffeur(request, chauffeur_id):
    """
    Vue détaillée des activités d'un chauffeur spécifique
    
    Affiche toutes les activités d'un chauffeur avec possibilité de voir
    les détails complets et les performances.
    """
    # Vérifier que le chauffeur est accessible au superviseur
    chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
    chauffeur = get_object_or_404(Chauffeur, id=chauffeur_id, id__in=chauffeurs_accessibles.values_list('id', flat=True))
    
    # Récupération des activités du chauffeur
    prises = PriseCles.objects.filter(chauffeur=chauffeur).order_by('-date', '-heure_prise')
    remises = RemiseCles.objects.filter(chauffeur=chauffeur).order_by('-date', '-heure_remise')
    
    # Pagination pour les prises
    prises_paginator = Paginator(prises, 15)
    prises_page = request.GET.get('prises_page')
    prises_obj = prises_paginator.get_page(prises_page)
    
    # Pagination pour les remises
    remises_paginator = Paginator(remises, 15)
    remises_page = request.GET.get('remises_page')
    remises_obj = remises_paginator.get_page(remises_page)
    
    # Statistiques du chauffeur
    total_prises = prises.count()
    total_remises = remises.count()
    
    # Calcul des performances
    performances = []
    for remise in remises:
        try:
            prise = PriseCles.objects.get(chauffeur=chauffeur, date=remise.date)
            pourcentage = (remise.recette_realisee / prise.objectif_recette) * 100
            performances.append({
                'date': remise.date,
                'objectif': prise.objectif_recette,
                'realise': remise.recette_realisee,
                'pourcentage': pourcentage,
                'statut': 'success' if pourcentage >= 100 else 'warning' if pourcentage >= 90 else 'danger'
            })
        except PriseCles.DoesNotExist:
            performances.append({
                'date': remise.date,
                'objectif': 0,
                'realise': remise.recette_realisee,
                'pourcentage': 0,
                'statut': 'info'
            })
    
    # Statistiques globales
    recettes_totales = remises.aggregate(total=Sum('recette_realisee'))['total'] or 0
    objectifs_totaux = prises.aggregate(total=Sum('objectif_recette'))['total'] or 0
    performance_moyenne = (recettes_totales / objectifs_totaux * 100) if objectifs_totaux > 0 else 0
    
    context = {
        'chauffeur': chauffeur,
        'prises': prises_obj,
        'remises': remises_obj,
        'prises_paginator': prises_paginator,
        'remises_paginator': remises_paginator,
        'performances': performances,
        'total_prises': total_prises,
        'total_remises': total_remises,
        'recettes_totales': recettes_totales,
        'objectifs_totaux': objectifs_totaux,
        'performance_moyenne': performance_moyenne,
    }
    
    return render(request, 'admin_dashboard/activites_chauffeur.html', context)


@supervisor_required
def exporter_activite_chauffeur_pdf(request, chauffeur_id):
    """
    Vue pour exporter les activités d'un chauffeur en PDF
    
    Génère un rapport PDF mensuel complet des activités d'un chauffeur avec
    les jours travaillés, recettes réalisées, performances et totaux par semaine.
    """
    # Vérifier que le chauffeur est accessible au superviseur
    chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
    chauffeur = get_object_or_404(Chauffeur, id=chauffeur_id, id__in=chauffeurs_accessibles.values_list('id', flat=True))
    
    # Récupération des paramètres de filtrage (par défaut: mois en cours)
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Si aucune date n'est spécifiée, utiliser le mois en cours
    if not date_debut or not date_fin:
        today = timezone.now().date()
        # Premier jour du mois
        date_debut = today.replace(day=1)
        # Dernier jour du mois
        if today.month == 12:
            date_fin = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            date_fin = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    
    # Conversion des dates
    if isinstance(date_debut, str):
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
    if isinstance(date_fin, str):
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
    
    # Récupération des activités du chauffeur pour le mois
    prises = PriseCles.objects.filter(
        chauffeur=chauffeur,
        date__gte=date_debut,
        date__lte=date_fin
    ).order_by('date', 'heure_prise')
    
    remises = RemiseCles.objects.filter(
        chauffeur=chauffeur,
        date__gte=date_debut,
        date__lte=date_fin
    ).order_by('date', 'heure_remise')
    
    # Calcul des statistiques générales
    total_prises = prises.count()
    total_remises = remises.count()
    recettes_totales = remises.aggregate(total=Sum('recette_realisee'))['total'] or 0
    objectifs_totaux = prises.aggregate(total=Sum('objectif_recette'))['total'] or 0
    performance_moyenne = (recettes_totales / objectifs_totaux * 100) if objectifs_totaux > 0 else 0
    
    # Calcul des performances par jour avec détails
    performances_journalieres = []
    jours_travailles = set()
    
    for remise in remises:
        jours_travailles.add(remise.date)
        try:
            prise = PriseCles.objects.get(chauffeur=chauffeur, date=remise.date)
            pourcentage = (remise.recette_realisee / prise.objectif_recette) * 100
            performances_journalieres.append({
                'date': remise.date,
                'jour_semaine': remise.date.strftime('%A'),
                'objectif': prise.objectif_recette,
                'realise': remise.recette_realisee,
                'pourcentage': pourcentage,
                'statut': 'success' if pourcentage >= 100 else 'warning' if pourcentage >= 90 else 'danger',
                'heure_prise': prise.heure_prise,
                'heure_remise': remise.heure_remise,
                'plein_carburant': remise.plein_carburant,
                'probleme_mecanique': remise.probleme_mecanique
            })
        except PriseCles.DoesNotExist:
            performances_journalieres.append({
                'date': remise.date,
                'jour_semaine': remise.date.strftime('%A'),
                'objectif': 0,
                'realise': remise.recette_realisee,
                'pourcentage': 0,
                'statut': 'info',
                'heure_prise': None,
                'heure_remise': remise.heure_remise,
                'plein_carburant': remise.plein_carburant,
                'probleme_mecanique': remise.probleme_mecanique
            })
    
    # Calcul des totaux par semaine
    totaux_semaines = []
    semaine_courante = None
    recette_semaine = 0
    objectif_semaine = 0
    jours_semaine = 0
    
    for perf in performances_journalieres:
        # Calcul du numéro de semaine
        semaine_num = perf['date'].isocalendar()[1]
        
        if semaine_courante is None:
            semaine_courante = semaine_num
        
        if semaine_num != semaine_courante:
            # Finaliser la semaine précédente
            totaux_semaines.append({
                'semaine': semaine_courante,
                'recette': recette_semaine,
                'objectif': objectif_semaine,
                'jours': jours_semaine,
                'performance': (recette_semaine / objectif_semaine * 100) if objectif_semaine > 0 else 0
            })
            # Commencer une nouvelle semaine
            semaine_courante = semaine_num
            recette_semaine = perf['realise']
            objectif_semaine = perf['objectif']
            jours_semaine = 1
        else:
            recette_semaine += perf['realise']
            objectif_semaine += perf['objectif']
            jours_semaine += 1
    
    # Ajouter la dernière semaine
    if semaine_courante is not None:
        totaux_semaines.append({
            'semaine': semaine_courante,
            'recette': recette_semaine,
            'objectif': objectif_semaine,
            'jours': jours_semaine,
            'performance': (recette_semaine / objectif_semaine * 100) if objectif_semaine > 0 else 0
        })
    
    # Préparation du contexte pour le PDF
    context = {
        'chauffeur': chauffeur,
        'prises': prises,
        'remises': remises,
        'performances_journalieres': performances_journalieres,
        'totaux_semaines': totaux_semaines,
        'total_prises': total_prises,
        'total_remises': total_remises,
        'recettes_totales': recettes_totales,
        'objectifs_totaux': objectifs_totaux,
        'performance_moyenne': performance_moyenne,
        'jours_travailles': len(jours_travailles),
        'date_debut': date_debut,
        'date_fin': date_fin,
        'date_generation': timezone.now(),
        'mois_nom': date_debut.strftime('%B %Y'),
        'superviseur_generateur': request.user,
    }
    
    # Génération du rapport en HTML pour impression/PDF
    # Solution simple sans dépendances externes - utilise le template HTML existant
    try:
        from django.template.loader import render_to_string
        from django.http import HttpResponse
        
        # Utiliser le template HTML existant avec des styles optimisés pour l'impression
        html_string = render_to_string('admin_dashboard/rapport_mensuel_chauffeur_pdf.html', context)
        
        # Retourner le HTML directement pour affichage dans le navigateur
        # L'utilisateur pourra l'imprimer en PDF via le bouton intégré (Ctrl+P -> Enregistrer en PDF)
        response = HttpResponse(html_string, content_type='text/html; charset=utf-8')
        
        return response
        
    except Exception as e:
        messages.error(request, f'Erreur lors de la génération du rapport: {str(e)}')
        return redirect('admin_dashboard:activites_chauffeur', chauffeur_id=chauffeur_id)


@supervisor_required
def gestion_demandes_modification(request):
    """
    Vue pour la gestion des demandes de modification d'activité
    
    Permet aux administrateurs de voir toutes les demandes de modification
    et de les approuver ou rejeter avec commentaires.
    """
    # Récupération des paramètres de filtrage
    statut = request.GET.get('statut', 'en_attente')
    chauffeur_id = request.GET.get('chauffeur')
    
    # Récupération des demandes avec filtres (selon les permissions utilisateur)
    chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
    demandes = DemandeModification.objects.select_related('chauffeur', 'admin_traite').filter(
        chauffeur__in=chauffeurs_accessibles
    )
    
    if statut:
        demandes = demandes.filter(statut=statut)
    
    if chauffeur_id:
        demandes = demandes.filter(chauffeur_id=chauffeur_id)
    
    # Tri par date de création décroissante
    demandes = demandes.order_by('-date_creation')
    
    # Pagination
    paginator = Paginator(demandes, 15)  # 15 éléments par page
    page_number = request.GET.get('page')
    demandes_obj = paginator.get_page(page_number)
    
    # Liste des chauffeurs pour le filtre (filtrée selon les permissions)
    chauffeurs = chauffeurs_accessibles.filter(actif=True).order_by('nom', 'prenom')
    
    # Statistiques
    total_demandes = demandes.count()
    demandes_en_attente = DemandeModification.objects.filter(statut='en_attente').count()
    demandes_approuvees = DemandeModification.objects.filter(statut='approuvee').count()
    demandes_rejetees = DemandeModification.objects.filter(statut='rejetee').count()
    
    context = {
        'demandes': demandes_obj,
        'paginator': paginator,
        'chauffeurs': chauffeurs,
        'total_demandes': total_demandes,
        'demandes_en_attente': demandes_en_attente,
        'demandes_approuvees': demandes_approuvees,
        'demandes_rejetees': demandes_rejetees,
        'filtres': {
            'statut': statut,
            'chauffeur_id': chauffeur_id,
        }
    }
    
    return render(request, 'admin_dashboard/gestion_demandes_modification.html', context)


@supervisor_required
def traiter_demande_modification(request, demande_id):
    """
    Vue pour traiter une demande de modification spécifique
    
    Permet à l'administrateur d'approuver ou rejeter une demande
    avec commentaires et d'appliquer les modifications si approuvée.
    """
    # Vérifier que la demande concerne un chauffeur accessible
    chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
    demande = get_object_or_404(DemandeModification, id=demande_id, chauffeur__in=chauffeurs_accessibles)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        commentaire = request.POST.get('commentaire', '')
        
        if action in ['approuver', 'rejeter']:
            # Traitement de la demande
            demande.traiter(
                admin=request.user,
                approuvee=(action == 'approuver'),
                commentaire=commentaire
            )
            
            # Si approuvée, appliquer les modifications
            if action == 'approuver':
                try:
                    if demande.type_activite == 'prise':
                        activite = PriseCles.objects.get(
                            chauffeur=demande.chauffeur, 
                            date=demande.date_activite
                        )
                    else:
                        activite = RemiseCles.objects.get(
                            chauffeur=demande.chauffeur, 
                            date=demande.date_activite
                        )
                    
                    # Application des nouvelles données
                    nouvelles_donnees = demande.nouvelles_donnees
                    for champ, valeur in nouvelles_donnees.items():
                        if hasattr(activite, champ):
                            setattr(activite, champ, valeur)
                    
                    activite.save()
                    
                    # Créer une panne si un problème mécanique est signalé dans les nouvelles données
                    nouveau_probleme = nouvelles_donnees.get('probleme_mecanique', '')
                    if nouveau_probleme and nouveau_probleme != 'Aucun':
                        from activities.models import Panne
                        Panne.objects.create(
                            chauffeur=demande.chauffeur,
                            description=nouveau_probleme,
                            severite='moderee',  # Par défaut
                            statut='signalee'  # Statut par défaut
                        )
                    
                    messages.success(request, f'Demande approuvée et modifications appliquées avec succès.')
                except Exception as e:
                    messages.error(request, f'Erreur lors de l\'application des modifications: {str(e)}')
            else:
                messages.success(request, 'Demande rejetée avec succès.')
            
            return redirect('admin_dashboard:gestion_demandes_modification')
    
    # Récupération de l'activité concernée pour affichage
    activite_actuelle = None
    try:
        if demande.type_activite == 'prise':
            activite_actuelle = PriseCles.objects.get(
                chauffeur=demande.chauffeur, 
                date=demande.date_activite
            )
        else:
            activite_actuelle = RemiseCles.objects.get(
                chauffeur=demande.chauffeur, 
                date=demande.date_activite
            )
    except (PriseCles.DoesNotExist, RemiseCles.DoesNotExist):
        pass
    
    context = {
        'demande': demande,
        'activite_actuelle': activite_actuelle,
    }
    
    return render(request, 'admin_dashboard/traiter_demande_modification.html', context)


@supervisor_required
def supprimer_activite(request, activite_id, type_activite):
    """
    Vue pour supprimer une activité spécifique (prise ou remise de clés)
    """
    # Vérifier les permissions d'accès
    chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
    
    try:
        if type_activite == 'prise':
            activite = get_object_or_404(PriseCles, id=activite_id, chauffeur__in=chauffeurs_accessibles)
        else:
            activite = get_object_or_404(RemiseCles, id=activite_id, chauffeur__in=chauffeurs_accessibles)
        
        chauffeur_nom = activite.chauffeur.nom_complet
        activite.delete()
        
        messages.success(request, f'Activité de {chauffeur_nom} supprimée avec succès.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('admin_dashboard:gestion_activites')


@supervisor_required
def supprimer_toutes_activites(request):
    """
    Vue pour supprimer toutes les activités (prises et remises de clés)
    """
    if request.method == 'POST':
        try:
            # Supprimer seulement les activités des chauffeurs accessibles
            chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
            
            # Supprimer les prises de clés
            prises_count = PriseCles.objects.filter(chauffeur__in=chauffeurs_accessibles).count()
            PriseCles.objects.filter(chauffeur__in=chauffeurs_accessibles).delete()
            
            # Supprimer les remises de clés
            remises_count = RemiseCles.objects.filter(chauffeur__in=chauffeurs_accessibles).count()
            RemiseCles.objects.filter(chauffeur__in=chauffeurs_accessibles).delete()
            
            messages.success(request, f'Toutes les activités ont été supprimées ({prises_count} prises, {remises_count} remises).')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('admin_dashboard:gestion_activites')


@supervisor_required
def supprimer_demande_modification(request, demande_id):
    """
    Vue pour supprimer une demande de modification spécifique
    """
    try:
        # Vérifier que la demande concerne un chauffeur accessible
        chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
        demande = get_object_or_404(DemandeModification, id=demande_id, chauffeur__in=chauffeurs_accessibles)
        chauffeur_nom = demande.chauffeur.nom_complet
        demande.delete()
        
        messages.success(request, f'Demande de modification de {chauffeur_nom} supprimée avec succès.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('admin_dashboard:gestion_demandes_modification')


@supervisor_required
def reinitialiser_demandes_modification(request):
    """
    Vue pour réinitialiser toutes les demandes de modification
    """
    if request.method == 'POST':
        try:
            # Supprimer seulement les demandes des chauffeurs accessibles
            chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
            demandes_count = DemandeModification.objects.filter(chauffeur__in=chauffeurs_accessibles).count()
            DemandeModification.objects.filter(chauffeur__in=chauffeurs_accessibles).delete()
            
            messages.success(request, f'Toutes les demandes de modification ont été supprimées ({demandes_count} demandes).')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('admin_dashboard:gestion_demandes_modification')


@supervisor_required
def supprimer_panne(request, panne_id):
    """
    Vue pour supprimer une panne spécifique
    """
    try:
        panne = get_object_or_404(Panne, id=panne_id)
        chauffeur_nom = panne.chauffeur.nom_complet
        panne.delete()
        
        messages.success(request, f'Panne de {chauffeur_nom} supprimée avec succès.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('admin_dashboard:dashboard_admin')


@supervisor_required
def supprimer_toutes_pannes(request):
    """
    Vue pour supprimer toutes les pannes
    """
    if request.method == 'POST':
        try:
            # Supprimer seulement les pannes des chauffeurs accessibles
            chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
            pannes_count = Panne.objects.filter(chauffeur__in=chauffeurs_accessibles).count()
            Panne.objects.filter(chauffeur__in=chauffeurs_accessibles).delete()
            
            messages.success(request, f'Toutes les pannes ont été supprimées ({pannes_count} pannes).')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('admin_dashboard:dashboard_admin')


@supervisor_required
def exporter_excel(request):
    """
    Export des données de recettes et performances en fichier Excel
    
    Génère un fichier Excel contenant toutes les données selon les filtres :
    - Période : jour, semaine, mois, année
    - Chauffeur : spécifique ou tous
    """
    # Vérifier si openpyxl est disponible
    if not OPENPYXL_AVAILABLE:
        messages.error(request, "Le module openpyxl n'est pas installé. Veuillez installer openpyxl pour utiliser cette fonctionnalité.")
        return redirect('admin_dashboard:statistiques_recettes')
    
    try:
        # Récupération des paramètres de filtre
        periode = request.GET.get('periode', 'mois')
        chauffeur_id = request.GET.get('chauffeur')
        
        # Calcul des dates selon la période
        aujourd_hui = timezone.now().date()
        
        if periode == 'jour':
            date_debut = aujourd_hui
            date_fin = aujourd_hui
        elif periode == 'semaine':
            # Lundi de cette semaine
            date_debut = aujourd_hui - timedelta(days=aujourd_hui.weekday())
            date_fin = date_debut + timedelta(days=6)
        elif periode == 'mois':
            date_debut = aujourd_hui.replace(day=1)
            if aujourd_hui.month == 12:
                date_fin = aujourd_hui.replace(year=aujourd_hui.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                date_fin = aujourd_hui.replace(month=aujourd_hui.month + 1, day=1) - timedelta(days=1)
        else:  # annee
            date_debut = aujourd_hui.replace(month=1, day=1)
            date_fin = aujourd_hui.replace(month=12, day=31)
        
        # Filtrage des données selon les permissions utilisateur
        chauffeurs_accessibles = get_chauffeurs_for_user(request.user)
        chauffeur_filter = Q(chauffeur__in=chauffeurs_accessibles)
        if chauffeur_id:
            chauffeur_filter &= Q(chauffeur_id=chauffeur_id)
        
        # Récupération des données
        prises = PriseCles.objects.filter(
            chauffeur_filter,
            date__range=[date_debut, date_fin]
        ).order_by('date', 'chauffeur__nom')
        
        remises = RemiseCles.objects.filter(
            chauffeur_filter,
            date__range=[date_debut, date_fin]
        ).order_by('date', 'chauffeur__nom')
        
        # Création du fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        # Titre limité à 31 caractères pour Excel
        ws.title = f"Recettes_{periode}_{date_debut.strftime('%m%d')}"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-têtes
        headers = [
            'Date', 'Chauffeur', 'Heure Début', 'Heure Fin', 
            'Objectif (FCFA)', 'Recette Réalisée (FCFA)', 'Performance (%)',
            'Carburant Plein', 'Problème Mécanique (Début)', 'Problème Mécanique (Fin)'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Données
        row = 2
        for prise in prises:
            # Trouver la remise correspondante
            remise = remises.filter(
                chauffeur=prise.chauffeur,
                date=prise.date
            ).first()
            
            # Calcul de la performance
            performance = 0
            if prise.objectif_recette and prise.objectif_recette > 0:
                recette_reelle = remise.recette_realisee if remise else 0
                performance = (recette_reelle / prise.objectif_recette) * 100
            
            # Données de la ligne
            data = [
                prise.date.strftime('%d/%m/%Y'),
                prise.chauffeur.nom_complet,
                prise.heure_prise.strftime('%H:%M') if prise.heure_prise else '',
                remise.heure_remise.strftime('%H:%M') if remise and remise.heure_remise else '',
                prise.objectif_recette or 0,
                remise.recette_realisee if remise else 0,
                round(performance, 1),
                'Oui' if prise.plein_carburant else 'Non',
                prise.probleme_mecanique or 'Aucun',
                remise.probleme_mecanique if remise else ''
            ]
            
            # Écriture des données
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [5, 6]:  # Colonnes monétaires
                    cell.number_format = '#,##0'
                elif col == 7:  # Performance
                    cell.number_format = '0.0'
            
            row += 1
        
        # Résumé
        ws.cell(row=row + 1, column=1, value="RÉSUMÉ").font = Font(bold=True)
        ws.cell(row=row + 2, column=1, value="Période:").font = Font(bold=True)
        ws.cell(row=row + 2, column=2, value=f"{date_debut.strftime('%d/%m/%Y')} - {date_fin.strftime('%d/%m/%Y')}")
        
        ws.cell(row=row + 3, column=1, value="Total Recettes:").font = Font(bold=True)
        total_recettes = sum(remise.recette_realisee for remise in remises if remise.recette_realisee)
        ws.cell(row=row + 3, column=2, value=total_recettes).number_format = '#,##0'
        
        ws.cell(row=row + 4, column=1, value="Total Objectifs:").font = Font(bold=True)
        total_objectifs = sum(prise.objectif_recette for prise in prises if prise.objectif_recette)
        ws.cell(row=row + 4, column=2, value=total_objectifs).number_format = '#,##0'
        
        ws.cell(row=row + 5, column=1, value="Performance Moyenne:").font = Font(bold=True)
        performance_moyenne = (total_recettes / total_objectifs * 100) if total_objectifs > 0 else 0
        ws.cell(row=row + 5, column=2, value=round(performance_moyenne, 1)).number_format = '0.0'
        
        # Ajustement des colonnes
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # Nom du fichier
        chauffeur_nom = ""
        if chauffeur_id:
            try:
                chauffeur = Chauffeur.objects.get(id=chauffeur_id)
                chauffeur_nom = f"_{chauffeur.nom}_{chauffeur.prenom}"
            except Chauffeur.DoesNotExist:
                pass
        
        filename = f"recettes_{periode}{chauffeur_nom}_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
        
        # Réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du fichier Excel : {str(e)}")
        return redirect('admin_dashboard:statistiques_recettes')


# =============================================================================
# GESTION DES SUPERVISEURS - Privilèges et permissions
# =============================================================================

@supervisor_required
def gestion_superviseurs(request):
    # Vérifier que l'utilisateur a les privilèges pour gérer les superviseurs
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Accès refusé. Vous n\'avez pas les privilèges nécessaires pour gérer les superviseurs.')
        return redirect('admin_dashboard:dashboard_admin')
    """
    Vue de gestion des superviseurs
    
    Permet à l'administrateur de :
    - Voir la liste des superviseurs
    - Ajouter/retirer des privilèges superviseur
    - Créer de nouveaux comptes superviseur
    
    Args:
        request: Objet HttpRequest de l'utilisateur
        
    Returns:
        HttpResponse: Template de gestion des superviseurs
    """
    from django.core.paginator import Paginator
    
    # Récupération du groupe Superviseurs
    try:
        superviseurs_group = Group.objects.get(name='Superviseurs')
        superviseurs = User.objects.filter(groups=superviseurs_group).order_by('username')
    except Group.DoesNotExist:
        superviseurs = User.objects.none()
        messages.warning(request, 'Le groupe "Superviseurs" n\'existe pas. Veuillez l\'initialiser.')
    
    # Pagination des superviseurs
    paginator = Paginator(superviseurs, 10)  # 10 superviseurs par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Récupération de tous les utilisateurs non-superviseurs avec pagination
    non_superviseurs = User.objects.exclude(groups=superviseurs_group).order_by('username')
    non_superviseurs_paginator = Paginator(non_superviseurs, 10)  # 10 utilisateurs par page
    non_superviseurs_page = request.GET.get('non_superviseurs_page')
    non_superviseurs_obj = non_superviseurs_paginator.get_page(non_superviseurs_page)
    
    context = {
        'superviseurs': page_obj,
        'page_obj': page_obj,
        'non_superviseurs': non_superviseurs_obj,
        'non_superviseurs_page_obj': non_superviseurs_obj,
        'superviseurs_count': superviseurs.count(),
    }
    
    return render(request, 'admin_dashboard/gestion_superviseurs.html', context)


@supervisor_required
def ajouter_superviseur(request, user_id):
    # Vérifier que l'utilisateur a les privilèges pour gérer les superviseurs
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Accès refusé. Vous n\'avez pas les privilèges nécessaires pour gérer les superviseurs.')
        return redirect('admin_dashboard:dashboard_admin')
    """
    Vue pour ajouter un utilisateur au groupe Superviseurs
    
    Args:
        request: Objet HttpRequest de l'utilisateur
        user_id: ID de l'utilisateur à promouvoir
        
    Returns:
        HttpResponse: Redirection vers la gestion des superviseurs
    """
    try:
        user = User.objects.get(id=user_id)
        superviseurs_group = Group.objects.get(name='Superviseurs')
        
        if not user.groups.filter(name='Superviseurs').exists():
            user.groups.add(superviseurs_group)
            messages.success(request, f'{user.username} a été ajouté aux superviseurs.')
        else:
            messages.warning(request, f'{user.username} est déjà superviseur.')
            
    except User.DoesNotExist:
        messages.error(request, 'Utilisateur non trouvé.')
    except Group.DoesNotExist:
        messages.error(request, 'Groupe "Superviseurs" non trouvé.')
    
    return redirect('admin_dashboard:gestion_superviseurs')


@supervisor_required
def retirer_superviseur(request, user_id):
    # Vérifier que l'utilisateur a les privilèges pour gérer les superviseurs
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Accès refusé. Vous n\'avez pas les privilèges nécessaires pour gérer les superviseurs.')
        return redirect('admin_dashboard:dashboard_admin')
    """
    Vue pour retirer un utilisateur du groupe Superviseurs
    
    Args:
        request: Objet HttpRequest de l'utilisateur
        user_id: ID de l'utilisateur à rétrograder
        
    Returns:
        HttpResponse: Redirection vers la gestion des superviseurs
    """
    try:
        user = User.objects.get(id=user_id)
        superviseurs_group = Group.objects.get(name='Superviseurs')
        
        if user.groups.filter(name='Superviseurs').exists():
            user.groups.remove(superviseurs_group)
            messages.success(request, f'{user.username} n\'est plus superviseur.')
        else:
            messages.warning(request, f'{user.username} n\'était pas superviseur.')
            
    except User.DoesNotExist:
        messages.error(request, 'Utilisateur non trouvé.')
    except Group.DoesNotExist:
        messages.error(request, 'Groupe "Superviseurs" non trouvé.')
    
    return redirect('admin_dashboard:gestion_superviseurs')


@supervisor_required
def creer_superviseur(request):
    # Vérifier que l'utilisateur a les privilèges pour gérer les superviseurs
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Accès refusé. Vous n\'avez pas les privilèges nécessaires pour gérer les superviseurs.')
        return redirect('admin_dashboard:dashboard_admin')
    """
    Vue pour créer un nouveau compte superviseur
    
    Args:
        request: Objet HttpRequest de l'utilisateur
        
    Returns:
        HttpResponse: Template de création ou redirection
    """
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        
        
        # Validation des champs obligatoires
        if not username or not password:
            messages.error(request, 'Nom d\'utilisateur et mot de passe sont obligatoires.')
            return render(request, 'admin_dashboard/creer_superviseur.html')
        
        # Vérification de l'unicité du nom d'utilisateur
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Ce nom d\'utilisateur existe déjà.')
            return render(request, 'admin_dashboard/creer_superviseur.html')
        
        try:
            # Création de l'utilisateur
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            
            # Ajout au groupe Superviseurs
            superviseurs_group, created = Group.objects.get_or_create(name='Superviseurs')
            user.groups.add(superviseurs_group)
            
            messages.success(request, f'Superviseur {username} créé avec succès.')
            return redirect('admin_dashboard:gestion_superviseurs')
            
        except Group.DoesNotExist:
            messages.error(request, 'Groupe "Superviseurs" non trouvé.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création : {str(e)}')
    
    return render(request, 'admin_dashboard/creer_superviseur.html')


@supervisor_required
def assigner_chauffeurs(request, superviseur_id):
    """
    Vue pour assigner des chauffeurs à un superviseur
    
    Args:
        request: Objet HttpRequest de l'utilisateur
        superviseur_id: ID du superviseur
        
    Returns:
        HttpResponse: Template d'assignation ou redirection
    """
    # Vérifier que l'utilisateur a les privilèges pour gérer les assignations
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Accès refusé. Vous n\'avez pas les privilèges nécessaires pour gérer les assignations.')
        return redirect('admin_dashboard:dashboard_admin')
    
    try:
        superviseur = User.objects.get(id=superviseur_id, groups__name='Superviseurs')
    except User.DoesNotExist:
        messages.error(request, 'Superviseur non trouvé.')
        return redirect('admin_dashboard:gestion_superviseurs')
    
    if request.method == 'POST':
        chauffeur_ids = request.POST.getlist('chauffeurs')
        
        
        try:
            # Approche sans transaction atomique pour éviter les rollbacks
            # 1. Supprimer TOUTES les assignations existantes pour ce superviseur
            old_assignments = AssignationSuperviseur.objects.filter(superviseur=superviseur)
            old_count = old_assignments.count()
            old_assignments.delete()
            
            # 2. Créer les nouvelles assignations une par une avec gestion d'erreur individuelle
            created_count = 0
            errors = []
            
            for chauffeur_id in chauffeur_ids:
                try:
                    chauffeur = Chauffeur.objects.get(id=chauffeur_id)
                    
                    # Utiliser get_or_create pour éviter les doublons
                    assignment, created = AssignationSuperviseur.objects.get_or_create(
                        chauffeur=chauffeur,
                        superviseur=superviseur,
                        defaults={
                            'assigne_par': request.user,
                            'actif': True
                        }
                    )
                    
                    if created:
                        created_count += 1
                    else:
                        # Réactiver si nécessaire
                        if not assignment.actif:
                            assignment.actif = True
                            assignment.assigne_par = request.user
                            assignment.save()
                    
                except Chauffeur.DoesNotExist:
                    error_msg = f"Chauffeur ID {chauffeur_id} non trouvé"
                    errors.append(error_msg)
                    continue
                except Exception as e:
                    error_msg = f"Erreur lors de l'assignation du chauffeur {chauffeur_id}: {str(e)}"
                    errors.append(error_msg)
                    continue
            
            # Messages de succès/erreur
            total_final = AssignationSuperviseur.objects.filter(superviseur=superviseur, actif=True).count()
            if total_final > 0:
                messages.success(request, f'Assignations mises à jour pour {superviseur.get_full_name() or superviseur.username}. {total_final} chauffeur(s) assigné(s).')
            else:
                messages.info(request, f'Toutes les assignations supprimées pour {superviseur.get_full_name() or superviseur.username}.')
            
            if errors:
                messages.warning(request, f'Quelques erreurs sont survenues: {"; ".join(errors[:3])}')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la mise à jour des assignations: {str(e)}')
        
        # TOUJOURS rediriger après POST (pattern PRG - Post-Redirect-Get)
        return redirect('admin_dashboard:gestion_superviseurs')
    
    # Vérifier que l'utilisateur a les privilèges pour cette fonctionnalité
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Vous n\'avez pas les privilèges nécessaires pour assigner des chauffeurs aux superviseurs.')
        return redirect('admin_dashboard:gestion_superviseurs')
    
    # Récupérer tous les chauffeurs (superuser a accès à tous)
    chauffeurs = Chauffeur.objects.filter(actif=True).order_by('nom', 'prenom')
    
    # Récupérer les chauffeurs déjà assignés
    chauffeurs_assignes = AssignationSuperviseur.get_chauffeurs_assignes(superviseur)
    
    context = {
        'superviseur': superviseur,
        'chauffeurs': chauffeurs,
        'chauffeurs_assignes': chauffeurs_assignes,
    }
    
    return render(request, 'admin_dashboard/assigner_chauffeurs.html', context)


@supervisor_required
def detail_superviseur(request, superviseur_id):
    """
    Vue pour afficher les détails d'un superviseur et ses chauffeurs assignés
    
    Args:
        request: Objet HttpRequest de l'utilisateur
        superviseur_id: ID du superviseur
        
    Returns:
        HttpResponse: Template de détail du superviseur
    """
    # Vérifier que l'utilisateur a les privilèges pour voir les détails des superviseurs
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Accès refusé. Vous n\'avez pas les privilèges nécessaires pour voir les détails des superviseurs.')
        return redirect('admin_dashboard:dashboard_admin')
    
    try:
        superviseur = User.objects.get(id=superviseur_id, groups__name='Superviseurs')
    except User.DoesNotExist:
        messages.error(request, 'Superviseur non trouvé.')
        return redirect('admin_dashboard:gestion_superviseurs')
    
    # Récupérer les chauffeurs assignés
    chauffeurs_assignes = AssignationSuperviseur.get_chauffeurs_assignes(superviseur)
    
    # Statistiques du superviseur
    stats = {
        'total_chauffeurs': chauffeurs_assignes.count(),
        'chauffeurs_actifs': chauffeurs_assignes.filter(actif=True).count(),
        'total_prises_mois': get_activites_for_user(superviseur, PriseCles, date__gte=date.today().replace(day=1)).count(),
        'total_remises_mois': get_activites_for_user(superviseur, RemiseCles, date__gte=date.today().replace(day=1)).count(),
        'recettes_mois': get_activites_for_user(superviseur, RemiseCles, date__gte=date.today().replace(day=1)).aggregate(total=Sum('recette_realisee'))['total'] or 0,
    }
    
    # Activités récentes des chauffeurs assignés
    activites_recentes = []
    prises_recentes = get_activites_for_user(superviseur, PriseCles).select_related('chauffeur').order_by('-date', '-heure_prise')[:10]
    remises_recentes = get_activites_for_user(superviseur, RemiseCles).select_related('chauffeur').order_by('-date', '-heure_remise')[:10]
    
    for prise in prises_recentes:
        activites_recentes.append({
            'type': 'Prise de clés',
            'chauffeur': prise.chauffeur.nom_complet,
            'date_heure': datetime.combine(prise.date, prise.heure_prise),
            'details': f"Objectif: {prise.objectif_recette} FCFA",
            'icon': 'bi-key text-primary'
        })
    
    for remise in remises_recentes:
        activites_recentes.append({
            'type': 'Remise de clés',
            'chauffeur': remise.chauffeur.nom_complet,
            'date_heure': datetime.combine(remise.date, remise.heure_remise),
            'details': f"Recette: {remise.recette_realisee} FCFA",
            'icon': 'bi-box-arrow-in-right text-success'
        })
    
    # Trier les activités récentes par date et heure
    activites_recentes.sort(key=lambda x: x['date_heure'], reverse=True)
    
    context = {
        'superviseur': superviseur,
        'chauffeurs_assignes': chauffeurs_assignes,
        'stats': stats,
        'activites_recentes': activites_recentes[:10],
    }
    
    return render(request, 'admin_dashboard/detail_superviseur.html', context)

    # Calcul des performances

    performances = []

    for remise in remises:

        try:

            prise = PriseCles.objects.get(chauffeur=chauffeur, date=remise.date)

            pourcentage = (remise.recette_realisee / prise.objectif_recette) * 100

            performances.append({

                'date': remise.date,

                'objectif': prise.objectif_recette,

                'realise': remise.recette_realisee,

                'pourcentage': pourcentage,

                'statut': 'success' if pourcentage >= 100 else 'warning' if pourcentage >= 90 else 'danger'

            })

        except PriseCles.DoesNotExist:

            performances.append({

                'date': remise.date,

                'objectif': 0,

                'realise': remise.recette_realisee,

                'pourcentage': 0,

                'statut': 'info'

            })

    

    # Statistiques globales

    recettes_totales = remises.aggregate(total=Sum('recette_realisee'))['total'] or 0

    objectifs_totaux = prises.aggregate(total=Sum('objectif_recette'))['total'] or 0

    performance_moyenne = (recettes_totales / objectifs_totaux * 100) if objectifs_totaux > 0 else 0

    

    context = {

        'chauffeur': chauffeur,

        'prises': prises_obj,

        'remises': remises_obj,

        'prises_paginator': prises_paginator,

        'remises_paginator': remises_paginator,

        'performances': performances,

        'total_prises': total_prises,

        'total_remises': total_remises,

        'recettes_totales': recettes_totales,

        'objectifs_totaux': objectifs_totaux,

        'performance_moyenne': performance_moyenne,

    }

    

    return render(request, 'admin_dashboard/activites_chauffeur.html', context)
